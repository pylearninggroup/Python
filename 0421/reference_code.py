# coding: utf-8
# scripts - login.py
# 2019/3/28 15:03



import pymysql
import openpyxl

# 表结构
# create table hello
# (
# 	name varchar(20) null,
# 	student_id int null,
# 	major varchar(20) null,
# 	score float null,
# 	address varchar(255) null
# );

#
con = pymysql.connect('127.0.0.1', 'root', 'root', charset='utf8mb4')
cur = con.cursor()

wb = openpyxl.load_workbook('sample.xlsx')
ws = wb.active


# 思路：把Excel的数据按照行读取
# 整理成[(),(),()...]或者[{},{},{}]的格式
# 然后使用executemany，不过需要注意内存问题、单条SQL语句最大长度、MySQL配置问题
# 或者读一行，写一行（速度稍慢，但是不用考虑以上问题）

data = []
# openpyxl的坐标是从1开始的，我们忽略表头，所以是2
for x in range(2, ws.max_row + 1):
    line = []
    for y in range(1, ws.max_column + 1):
        # 这里读取的就是一行的数据
        line.append(ws.cell(x, y).value)
    # 把每一行的结果（也是个列表）追加到data中，data就是我们需要的格式
    data.append(line)

cur.execute('USE abcd')
# 这里使用executemany
cur.executemany("INSERT INTO hello values (%s,%s,%s,%s,%s)", data)
con.commit()
con.close()
